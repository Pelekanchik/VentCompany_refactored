#!/usr/bin/env python3
"""Автоматичне виправлення типових помилок ruff для VentCompany."""
from pathlib import Path

ROOT = Path(".")


def fix_file(path: Path, replacements: list):
    text = path.read_text(encoding="utf-8")
    original = text
    for old, new in replacements:
        text = text.replace(old, new)
    if text != original:
        path.write_text(text, encoding="utf-8")
        print(f"  ✅ {path}")
    else:
        print(f"  ⏭️  {path} (без змін)")


print("🔧 Виправлення main_cli.py ...")
fix_file(
    ROOT / "main_cli.py",
    [
        ("calc_result = demo_calculations(project)", "demo_calculations(project)"),
        (
            'd = int(input("   Діаметр (мм): "))\n                l = float(input("   Довжина (м): "))',
            'd = int(input("   Діаметр (мм): "))\n                length = float(input("   Довжина (м): "))',
        ),
        (
            "result = calc.calculate_round_duct(d, l)",
            "result = calc.calculate_round_duct(d, length)",
        ),
        (
            'w = int(input("   Ширина (мм): "))\n                h = int(input("   Висота (мм): "))\n                l = float(input("   Довжина (м): "))',
            'w = int(input("   Ширина (мм): "))\n                h = int(input("   Висота (мм): "))\n                length = float(input("   Довжина (м): "))',
        ),
        (
            "result = calc.calculate_rectangular_duct(w, h, l)",
            "result = calc.calculate_rectangular_duct(w, h, length)",
        ),
    ],
)

print("🔧 Виправлення cad_editor.py ...")
fix_file(
    ROOT / "ventilation_company" / "cad_editor.py",
    [
        (
            "command=lambda l=layer_name: self.toggle_layer(l)",
            "command=lambda _layer=layer_name: self.toggle_layer(_layer)",
        ),
        (
            "isinstance(e, CADLine) or isinstance(e, CADRectangle)",
            "isinstance(e, (CADLine, CADRectangle))",
        ),
    ],
)

print("🔧 Виправлення camduct_editor.py ...")
fix_file(
    ROOT / "ventilation_company" / "camduct_editor.py",
    [
        (
            "for i, (key, label, default) in enumerate(fields):",
            "for _i, (key, label, default) in enumerate(fields):",
        ),
        (
            'l = float(sizes.get("length", 1000))\n            unfold_w = 2 * (w + h) + 6  # периметр + 6 мм на фальці\n            unfold_h = l',
            'length = float(sizes.get("length", 1000))\n            unfold_w = 2 * (w + h) + 6  # периметр + 6 мм на фальці\n            unfold_h = length',
        ),
        (
            'l = float(sizes.get("length", 1000))\n            unfold_w = math.pi * (d + 6)  # π*(d+6) — з урахуванням фальця\n            unfold_h = l',
            'length = float(sizes.get("length", 1000))\n            unfold_w = math.pi * (d + 6)  # π*(d+6) — з урахуванням фальця\n            unfold_h = length',
        ),
        (
            'l = float(sizes.get("length", 300))\n            unfold_w = math.pi * ((d1 + d2) / 2 + 6)\n            unfold_h = l * 1.25',
            'length = float(sizes.get("length", 300))\n            unfold_w = math.pi * ((d1 + d2) / 2 + 6)\n            unfold_h = length * 1.25',
        ),
        (
            'l = float(sizes.get("length", 300))\n            unfold_w = 2 * ((w1 + h1 + w2 + h2) / 2 + 6)\n            unfold_h = l * 1.3',
            'length = float(sizes.get("length", 300))\n            unfold_w = 2 * ((w1 + h1 + w2 + h2) / 2 + 6)\n            unfold_h = length * 1.3',
        ),
    ],
)

print("🔧 Виправлення calc_repo.py ...")
calc_repo = ROOT / "ventilation_company" / "database" / "repositories" / "calc_repo.py"
text = calc_repo.read_text(encoding="utf-8")
lines = text.splitlines()
new_lines = []
count = 0
for line in lines:
    if "def get_calculations_by_period(" in line:
        count += 1
        if count == 2:
            line = line.replace("get_calculations_by_period", "get_calculations_by_date_range")
    new_lines.append(line)
calc_repo.write_text("\n".join(new_lines), encoding="utf-8")
print("  ✅ calc_repo.py (перейменовано другу функцію)")

print("🔧 Виправлення settings_repo.py ...")
fix_file(
    ROOT / "ventilation_company" / "database" / "repositories" / "settings_repo.py",
    [
        ("except:\n                pass", "except Exception:\n                pass"),
    ],
)

print("🔧 Виправлення db_core.py (E402) ...")
db_core = ROOT / "ventilation_company" / "db_core.py"
text = db_core.read_text(encoding="utf-8")
for imp in ["import json\n", "from math import pi\n"]:
    text = text.replace(imp, "")
if "import json" not in text.split("# ==========")[0]:
    text = text.replace(
        "from ventilation_company.config import DB_PATH\n",
        "from ventilation_company.config import DB_PATH\nimport json\nfrom math import pi\n",
    )
db_core.write_text(text, encoding="utf-8")
print("  ✅ db_core.py")

print("🔧 Виправлення detail_calculator.py ...")
dc = ROOT / "ventilation_company" / "detail_calculator.py"
text = dc.read_text(encoding="utf-8")
old_import = """try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True"""
new_import = """try:
    import openpyxl

    HAS_OPENPYXL = True"""
text = text.replace(old_import, new_import)
text = text.replace("except:\n            pass", "except Exception:\n            pass")
text = text.replace("* flange_waste", "* flange_waste_factor")
dc.write_text(text, encoding="utf-8")
print("  ✅ detail_calculator.py")

print("🔧 Виправлення gui/__init__.py ...")
fix_file(
    ROOT / "ventilation_company" / "gui" / "__init__.py",
    [
        ("app = VentilationApp(root)", "VentilationApp(root)"),
    ],
)

print("🔧 Виправлення gui/base_tab.py ...")
fix_file(
    ROOT / "ventilation_company" / "gui" / "base_tab.py",
    [
        (
            "except ValueError:\n              raise ValueError(",
            "except ValueError as err:\n              raise ValueError(",
        ),
    ],
)
bt = ROOT / "ventilation_company" / "gui" / "base_tab.py"
text = bt.read_text(encoding="utf-8")
if "from err" not in text:
    text = text.replace(
        ")\n            )\n        return float(raw)",
        ") from err\n            )\n        return float(raw)",
    )
    bt.write_text(text, encoding="utf-8")
    print("  ✅ gui/base_tab.py (додано from err)")

print("🔧 Виправлення gui/main_window.py ...")
fix_file(
    ROOT / "ventilation_company" / "gui" / "main_window.py",
    [
        ("for theme_name in self.themes.keys():", "for theme_name in self.themes:"),
        ("for name, frame in self.tabs.items():", "for _name, frame in self.tabs.items():"),
        ("TabClass = self._tab_classes[tab_name]", "tab_class = self._tab_classes[tab_name]"),
        (
            "self.tabs[tab_name] = TabClass(self.tabs_container, self)",
            "self.tabs[tab_name] = tab_class(self.tabs_container, self)",
        ),
    ],
)

print("🔧 Виправлення gui/price_list_tab.py ...")
fix_file(
    ROOT / "ventilation_company" / "gui" / "price_list_tab.py",
    [
        ("c = self.colors\n        OriginalPriceListTab", "OriginalPriceListTab"),
    ],
)

print("🔧 Виправлення price_list_tab.py ...")
plt = ROOT / "ventilation_company" / "price_list_tab.py"
text = plt.read_text(encoding="utf-8")
text = text.replace("raise SystemExit(1)", "raise SystemExit(1) from None")
if "from ventilation_company.models.product import Product" in text:
    lines = text.splitlines()
    import_line = None
    other_lines = []
    for line in lines:
        if line.strip() == "from ventilation_company.models.product import Product":
            import_line = line
        else:
            other_lines.append(line)
    if import_line:
        for i, line in enumerate(other_lines):
            if line.startswith("import ") or line.startswith("from "):
                last_import = i
        other_lines.insert(last_import + 1, import_line)
        text = "\n".join(other_lines)
text = text.replace(
    "for key, label, numeric, default in text_fields:",
    "for key, label, _numeric, default in text_fields:",
)
text = text.replace("btn_cfg = dict(", "btn_cfg = {")
text = text.replace(
    '        )\n\n        tk.Button(toolbar, text="➕ Додати", command=self._add, **btn_cfg)',
    '        }\n\n        tk.Button(toolbar, text="➕ Додати", command=self._add, **btn_cfg)',
)
text = text.replace(
    "for col, (attr, typ, editable, is_combo, width) in enumerate(self.col_map):",
    "for _col, (_attr, _typ, _editable, _is_combo, width) in enumerate(self.col_map):",
)
text = text.replace(
    "for col, (attr, _, _, _, _) in enumerate(self.col_map):",
    "for _col, (attr, _, _, _, _) in enumerate(self.col_map):",
)
text = text.replace(
    """        if self._edit_is_combo:
            raw = self._edit_widget.get()
        else:
            raw = self._edit_widget.get().strip()""",
    "        raw = self._edit_widget.get() if self._edit_is_combo else self._edit_widget.get().strip()",
)
plt.write_text(text, encoding="utf-8")
print("  ✅ price_list_tab.py")

print("🔧 Виправлення services/calculator_service.py ...")
cs = ROOT / "ventilation_company" / "services" / "calculator_service.py"
text = cs.read_text(encoding="utf-8")
text = text.replace(
    """        if labor_mode == "m2":
            labor_cost = area * labor_rate
        else:
            labor_cost = labor_norm * labor_rate""",
    '        labor_cost = area * labor_rate if labor_mode == "m2" else labor_norm * labor_rate',
)
cs.write_text(text, encoding="utf-8")
print("  ✅ calculator_service.py")

print("🔧 Виправлення services/export_service.py ...")
fix_file(
    ROOT / "ventilation_company" / "services" / "export_service.py",
    [
        (
            "from openpyxl.styles import Alignment, Border, Font, PatternFill, Side",
            "from openpyxl.styles import Alignment, Font, PatternFill",
        ),
        ("except:\n                    pass", "except Exception:\n                    pass"),
    ],
)

print("\n🎉 Готово! Перевірте: ruff check .")

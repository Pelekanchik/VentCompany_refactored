#!/usr/bin/env python3
"""
Прайс-лист вентиляційних виробів — ВКЛАДКА
(адаптовано для інтеграції в VentilationApp)
"""

import csv
import json
import os
import platform
import shutil
import subprocess
import tempfile
import uuid
import webbrowser
from datetime import datetime
from typing import Any

from ventilation_company.models.product import Product

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    print("❌ tkinter не знайдено.")
    raise SystemExit(1) from None

DATA_FILE = "ventilation_price_list.json"
EXPORT_DIR = "exports"
BACKUP_DIR = "backups"
MATERIALS = ["цинк", "нержавійка"]
CATEGORIES = [
    "Вентилятор",
    "Труба прямокутна",
    "Труба кругла",
    "Фасонка",
    "Заслінка",
    "Решітка",
    "Клапан",
    "Комплектуюче",
    "Інше",
]

HAS_OPENPYXL = False
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    pass


def open_folder_crossplatform(path: str):
    abs_path = os.path.abspath(path)
    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(abs_path)
        elif system == "Darwin":
            subprocess.call(["open", abs_path])
        else:
            subprocess.call(["xdg-open", abs_path])
    except Exception as e:
        messagebox.showerror("Помилка", "Не вдалося відкрити папку: " + str(e))


def generate_id() -> str:
    return str(uuid.uuid4())[:8]


class PriceList:
    def __init__(self, filepath: str = DATA_FILE):
        self.filepath = filepath
        self.products: list[Product] = []
        self._undo_stack: list[list[dict]] = []
        self._redo_stack: list[list[dict]] = []
        self._max_undo = 20
        self._load()

    def _snapshot(self) -> list[dict]:
        return [p.to_dict() for p in self.products]

    def _restore(self, snapshot: list[dict]):
        self.products = [Product.from_dict(d) for d in snapshot]

    def _push_undo(self):
        self._undo_stack.append(self._snapshot())
        if len(self._undo_stack) > self._max_undo:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def undo(self) -> bool:
        if not self._undo_stack:
            return False
        current = self._snapshot()
        snapshot = self._undo_stack.pop()
        self._redo_stack.append(current)
        self._restore(snapshot)
        return True

    def redo(self) -> bool:
        if not self._redo_stack:
            return False
        current = self._snapshot()
        snapshot = self._redo_stack.pop()
        self._undo_stack.append(current)
        self._restore(snapshot)
        return True

    def _load(self) -> None:
        if not os.path.exists(self.filepath):
            return
        try:
            with open(self.filepath, encoding="utf-8") as f:
                raw = json.load(f)
            for item in raw:
                self.products.append(Product.from_dict(item))
        except Exception as e:
            print(f"Помилка завантаження: {e}")

    def save(self) -> None:
        self._backup()
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump([p.to_dict() for p in self.products], f, ensure_ascii=False, indent=2)

    def _backup(self):
        if not os.path.exists(self.filepath):
            return
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)
        timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")
        backup_path = os.path.join(BACKUP_DIR, f"backup_{timestamp}.json")
        try:
            shutil.copy2(self.filepath, backup_path)
            backups = sorted(
                [
                    f
                    for f in os.listdir(BACKUP_DIR)
                    if f.startswith("backup_") and f.endswith(".json")
                ]
            )
            for old in backups[:-10]:
                os.remove(os.path.join(BACKUP_DIR, old))
        except Exception:
            pass

    def add(self, product: Product) -> Product:
        self._push_undo()
        self.products.append(product)
        return product

    def bulk_add(self, products: list[Product]):
        self._push_undo()
        self.products.extend(products)

    def get_by_id(self, product_id: str) -> Product | None:
        for p in self.products:
            if p.id == product_id:
                return p
        return None

    def delete(self, product_id: str) -> bool:
        product = self.get_by_id(product_id)
        if product:
            self._push_undo()
            self.products.remove(product)
            return True
        return False

    def total_sum(self) -> float:
        return round(sum(p.total_price for p in self.products), 2)

    def bulk_edit_price(self, ids: list[str], new_price: float) -> int:
        self._push_undo()
        count = 0
        for p in self.products:
            if p.id in ids:
                old = p.price_per_unit
                p.price_per_unit = new_price
                p.record_price_change(old)
                count += 1
        return count

    def bulk_edit_price_percent(self, ids: list[str], percent: float) -> int:
        self._push_undo()
        count = 0
        for p in self.products:
            if p.id in ids:
                old = p.price_per_unit
                p.price_per_unit = round(p.price_per_unit * (1 + percent / 100), 2)
                p.record_price_change(old)
                count += 1
        return count

    def bulk_edit_quantity(self, ids: list[str], new_qty: float) -> int:
        self._push_undo()
        count = 0
        for p in self.products:
            if p.id in ids:
                p.quantity = new_qty
                count += 1
        return count

    def bulk_edit_material(self, ids: list[str], new_material: str) -> int:
        self._push_undo()
        count = 0
        for p in self.products:
            if p.id in ids:
                p.material = new_material
                count += 1
        return count

    def get_statistics(self) -> dict[str, Any]:
        if not self.products:
            return {}
        stats = {
            "total_products": len(self.products),
            "total_sum": self.total_sum(),
            "by_material": {},
            "by_category": {},
            "avg_price": round(
                sum(p.price_per_unit for p in self.products) / len(self.products), 2
            ),
            "total_quantity": sum(p.quantity for p in self.products),
        }
        for p in self.products:
            stats["by_material"][p.material] = (
                stats["by_material"].get(p.material, 0) + p.total_price
            )
            stats["by_category"][p.category] = (
                stats["by_category"].get(p.category, 0) + p.total_price
            )
        return stats


def generate_print_html(
    products: list[Product], grand_total: float, title: str = "Прайс-лист"
) -> str:
    rows = []
    for p in products:
        rows.append(
            f"""
        <tr>
            <td style="text-align:center;white-space:nowrap;">{p.date_only}</td>
            <td><b>{p.name}</b></td>
            <td style="text-align:center;">{p.category}</td>
            <td style="text-align:center;">{p.length:g}</td>
            <td style="text-align:center;">{p.width:g}</td>
            <td style="text-align:center;">{p.height:g}</td>
            <td style="text-align:center;">{p.diameter:g}</td>
            <td>{p.material}</td>
            <td style="text-align:center;">{p.thickness:g}</td>
            <td style="text-align:center;">{p.quantity:g}</td>
            <td style="text-align:right;">{p.price_per_unit:,.2f}</td>
            <td style="text-align:right;font-weight:bold;color:#2e7d32;">{p.total_price:,.2f}</td>
        </tr>
        """
        )

    html = f"""<!DOCTYPE html>
<html lang="uk">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<style>
    @page {{ size: A4 landscape; margin: 8mm; }}
    @media print {{
        body {{ margin: 0; }}
        .no-print {{ display: none !important; }}
        table {{ box-shadow: none !important; }}
    }}
    body {{
        font-family: "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        background: #f5f5f5;
        margin: 0;
        padding: 15px;
        color: #333;
    }}
    .container {{
        max-width: 1500px;
        margin: 0 auto;
        background: #fff;
        padding: 20px;
        border-radius: 8px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }}
    h1 {{
        text-align: center;
        color: #1565c0;
        margin-bottom: 5px;
        font-size: 24px;
    }}
    .subtitle {{
        text-align: center;
        color: #666;
        font-size: 12px;
        margin-bottom: 15px;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
    }}
    th {{
        background: #1565c0;
        color: #fff;
        padding: 8px 5px;
        text-align: center;
        font-weight: 600;
    }}
    td {{
        padding: 8px 5px;
        border-bottom: 1px solid #e0e0e0;
        vertical-align: middle;
    }}
    tr:hover {{ background: #f5f5f5; }}
    .no-print {{
        text-align: center;
        margin-bottom: 15px;
    }}
    .btn {{
        display: inline-block;
        padding: 10px 24px;
        background: #1565c0;
        color: #fff;
        border: none;
        border-radius: 6px;
        font-size: 15px;
        cursor: pointer;
        text-decoration: none;
        margin: 0 5px;
    }}
    .btn:hover {{ background: #0d47a1; }}
    .btn-secondary {{ background: #757575; }}
    .btn-secondary:hover {{ background: #424242; }}
    .grand-total {{
        text-align: right;
        margin-top: 15px;
        padding: 12px 20px;
        background: #e3f2fd;
        border-radius: 6px;
        font-size: 18px;
        font-weight: 700;
        color: #0d47a1;
    }}
    .total-label {{
        text-align: right;
        margin-top: 8px;
        font-size: 13px;
        color: #666;
    }}
</style>
</head>
<body>
<div class="no-print">
    <button class="btn" onclick="window.print()">🖨️ Друкувати</button>
    <button class="btn btn-secondary" onclick="window.close()">❌ Закрити</button>
</div>
<div class="container">
    <h1>📋 {title}</h1>
    <div class="subtitle">Вентиляційні вироби та комплектуючі &nbsp;|&nbsp; Друк: {datetime.now().strftime("%d.%m.%Y %H:%M")}</div>
    <table>
        <thead>
            <tr>
                <th style="width:7%">Дата</th>
                <th style="width:16%">Назва</th>
                <th style="width:10%">Категорія</th>
                <th style="width:6%">Довжина</th>
                <th style="width:6%">Ширина</th>
                <th style="width:6%">Висота</th>
                <th style="width:6%">Діаметр</th>
                <th style="width:8%">Матеріал</th>
                <th style="width:6%">Товщина</th>
                <th style="width:5%">К-ть</th>
                <th style="width:9%;text-align:right">Ціна за шт</th>
                <th style="width:9%;text-align:right">Загальна</th>
            </tr>
        </thead>
        <tbody>
            {''.join(rows)}
        </tbody>
    </table>
    <div class="total-label">Всього позицій: {len(products)}</div>
    <div class="grand-total">💰 ЗАГАЛЬНА СУМА: {grand_total:,.2f} грн</div>
</div>
</body>
</html>"""
    return html


def export_to_excel(products: list[Product], grand_total: float, filepath: str) -> bool:
    if not HAS_OPENPYXL:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    total_font = Font(bold=True, size=12, color="0D47A1")
    total_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    headers = [
        "Дата",
        "Назва виробу",
        "Категорія",
        "Довжина, мм",
        "Ширина, мм",
        "Висота, мм",
        "Діаметр, мм",
        "Матеріал",
        "Товщина, мм",
        "К-ть",
        "Ціна за шт, грн",
        "Загальна, грн",
        "Примітки",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, p in enumerate(products, 2):
        data = [
            p.date_only,
            p.name,
            p.category,
            p.length,
            p.width,
            p.height,
            p.diameter,
            p.material,
            p.thickness,
            p.quantity,
            p.price_per_unit,
            p.total_price,
            p.notes,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col not in (2, 13) else "left", vertical="center"
            )
            if col in (11, 12):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    sum_row = len(products) + 2
    ws.cell(row=sum_row, column=1, value="ЗАГАЛЬНА СУМА:").font = total_font
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=11)
    total_cell = ws.cell(row=sum_row, column=12, value=grand_total)
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = thin_border

    col_widths = [12, 28, 14, 12, 12, 12, 12, 14, 14, 8, 16, 16, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(filepath)
    return True


class AddProductDialog(tk.Toplevel):
    def __init__(self, parent, product_to_clone: Product | None = None):
        super().__init__(parent)
        self.title("Новий виріб" if product_to_clone is None else "Дублювати виріб")
        self.geometry("420x600")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None

        now_str = datetime.now().strftime("%d.%m.%Y %H:%M")
        clone = product_to_clone

        tk.Label(self, text=self.title(), font=("Segoe UI", 14, "bold"), fg="#1565c0").pack(
            pady=(15, 5)
        )
        tk.Label(self, text=f"Дата додавання: {now_str}", font=("Segoe UI", 9), fg="#666").pack(
            pady=(0, 10)
        )

        text_fields = [
            ("name", "Назва виробу *", False, clone.name if clone else ""),
            ("length", "Довжина (мм)", True, str(clone.length) if clone and clone.length else ""),
            ("width", "Ширина (мм)", True, str(clone.width) if clone and clone.width else ""),
            ("height", "Висота (мм)", True, str(clone.height) if clone and clone.height else ""),
            (
                "diameter",
                "Діаметр (мм)",
                True,
                str(clone.diameter) if clone and clone.diameter else "",
            ),
            (
                "thickness",
                "Товщина матеріалу (мм)",
                True,
                str(clone.thickness) if clone and clone.thickness else "",
            ),
            (
                "price_per_unit",
                "Ціна за штуку (грн) *",
                True,
                str(clone.price_per_unit) if clone else "",
            ),
            ("quantity", "Кількість *", True, str(clone.quantity) if clone else "1"),
            ("notes", "Примітки", False, clone.notes if clone else ""),
        ]

        self.entries = {}
        for key, label, _numeric, default in text_fields:
            frm = tk.Frame(self)
            frm.pack(fill="x", padx=20, pady=3)
            tk.Label(frm, text=label + ":", font=("Segoe UI", 9), width=32, anchor="w").pack(
                side="left"
            )
            ent = tk.Entry(frm, font=("Segoe UI", 10), width=22)
            ent.insert(0, default)
            ent.pack(side="left")
            self.entries[key] = ent

        frm_cat = tk.Frame(self)
        frm_cat.pack(fill="x", padx=20, pady=3)
        tk.Label(frm_cat, text="Категорія:", font=("Segoe UI", 9), width=32, anchor="w").pack(
            side="left"
        )
        self.category_var = tk.StringVar(value=clone.category if clone else CATEGORIES[0])
        ttk.Combobox(
            frm_cat,
            textvariable=self.category_var,
            values=CATEGORIES,
            state="readonly",
            font=("Segoe UI", 10),
            width=20,
        ).pack(side="left")

        frm_mat = tk.Frame(self)
        frm_mat.pack(fill="x", padx=20, pady=3)
        tk.Label(frm_mat, text="Матеріал:", font=("Segoe UI", 9), width=32, anchor="w").pack(
            side="left"
        )
        self.material_var = tk.StringVar(
            value=clone.material if clone and clone.material in MATERIALS else MATERIALS[0]
        )
        ttk.Combobox(
            frm_mat,
            textvariable=self.material_var,
            values=MATERIALS,
            state="readonly",
            font=("Segoe UI", 10),
            width=20,
        ).pack(side="left")

        frm = tk.Frame(self)
        frm.pack(pady=20)
        tk.Button(
            frm,
            text="💾 Зберегти виріб",
            bg="#1565c0",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            padx=25,
            pady=5,
            command=self._save,
        ).pack(side="left", padx=5)
        tk.Button(frm, text="Скасувати", font=("Segoe UI", 10), padx=15, command=self.destroy).pack(
            side="left", padx=5
        )

        self.entries["name"].focus()
        self.wait_window(self)

    def _save(self):
        name = self.entries["name"].get().strip()
        if not name:
            messagebox.showwarning("Увага", "Назва виробу обов'язкова!")
            return
        try:
            price_per_unit = float(self.entries["price_per_unit"].get().strip() or 0)
            quantity = float(self.entries["quantity"].get().strip() or 1)
            length = float(self.entries["length"].get().strip() or 0)
            width = float(self.entries["width"].get().strip() or 0)
            height = float(self.entries["height"].get().strip() or 0)
            diameter = float(self.entries["diameter"].get().strip() or 0)
            thickness = float(self.entries["thickness"].get().strip() or 0)
        except ValueError:
            messagebox.showwarning("Увага", "Ціна, кількість та розміри мають бути числами!")
            return

        self.result = Product(
            product_id=generate_id(),
            date_added=datetime.now().strftime("%d.%m.%Y %H:%M"),
            name=name,
            price_per_unit=price_per_unit,
            quantity=quantity,
            length=length,
            width=width,
            height=height,
            diameter=diameter,
            material=self.material_var.get(),
            thickness=thickness,
            category=self.category_var.get(),
            notes=self.entries["notes"].get().strip(),
        )
        self.destroy()


class BulkPriceDialog(tk.Toplevel):
    def __init__(self, parent, count: int):
        super().__init__(parent)
        self.title("Масова зміна ціни")
        self.geometry("340x260")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None

        tk.Label(
            self, text=f"Вибрано виробів: {count}", font=("Segoe UI", 11, "bold"), fg="#1565c0"
        ).pack(pady=10)

        self.mode_var = tk.StringVar(value="abs")
        frm_mode = tk.Frame(self)
        frm_mode.pack(pady=5)
        tk.Radiobutton(
            frm_mode,
            text="Нова ціна за шт",
            variable=self.mode_var,
            value="abs",
            command=self._toggle,
        ).pack(side="left", padx=10)
        tk.Radiobutton(
            frm_mode, text="На відсоток", variable=self.mode_var, value="pct", command=self._toggle
        ).pack(side="left", padx=10)

        self.lbl = tk.Label(self, text="Нова ціна за шт (грн):", font=("Segoe UI", 10))
        self.lbl.pack(pady=(10, 0))
        self.ent = tk.Entry(self, font=("Segoe UI", 12), justify="center", width=20)
        self.ent.pack(pady=5)
        self.ent.focus()

        self.hint = tk.Label(self, text="", font=("Segoe UI", 9), fg="#666")
        self.hint.pack()

        frm = tk.Frame(self)
        frm.pack(pady=15)
        tk.Button(
            frm,
            text="Застосувати",
            bg="#1565c0",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
            command=self._ok,
        ).pack(side="left", padx=5)
        tk.Button(frm, text="Скасувати", command=self.destroy).pack(side="left", padx=5)

        self._toggle()
        self.wait_window(self)

    def _toggle(self):
        if self.mode_var.get() == "abs":
            self.lbl.config(text="Нова ціна за шт (грн):")
            self.hint.config(text="Ціна буде встановлена однакова для всіх")
        else:
            self.lbl.config(text="Відсоток зміни (%):")
            self.hint.config(text="+10 — підвищити на 10%,  -15 — знизити на 15%")

    def _ok(self):
        raw = self.ent.get().strip()
        if not raw:
            messagebox.showwarning("Увага", "Введіть значення!")
            return
        try:
            val = float(raw)
        except ValueError:
            messagebox.showwarning("Увага", "Введіть число!")
            return
        self.result = (self.mode_var.get(), val)
        self.destroy()


class StatisticsDialog(tk.Toplevel):
    def __init__(self, parent, stats: dict[str, Any]):
        super().__init__(parent)
        self.title("📊 Статистика")
        self.geometry("450x500")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        if not stats:
            tk.Label(self, text="Немає даних для статистики", font=("Segoe UI", 12)).pack(pady=20)
            tk.Button(self, text="Закрити", command=self.destroy).pack(pady=10)
            self.wait_window(self)
            return

        canvas = tk.Canvas(self)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas)

        scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        def add_section(title, data_dict):
            frm = tk.LabelFrame(
                scroll_frame,
                text=title,
                font=("Segoe UI", 10, "bold"),
                fg="#1565c0",
                padx=10,
                pady=10,
            )
            frm.pack(fill="x", pady=5, padx=5)
            if not data_dict:
                tk.Label(frm, text="—", font=("Segoe UI", 9), fg="#666").pack()
                return
            for key, val in sorted(data_dict.items(), key=lambda x: -x[1]):
                row = tk.Frame(frm)
                row.pack(fill="x", pady=2)
                tk.Label(row, text=str(key), font=("Segoe UI", 9), anchor="w").pack(side="left")
                tk.Label(
                    row,
                    text=f"{val:,.2f} грн",
                    font=("Segoe UI", 9, "bold"),
                    fg="#2e7d32",
                    anchor="e",
                ).pack(side="right")

        tk.Label(
            scroll_frame, text="📊 Загальна статистика", font=("Segoe UI", 14, "bold"), fg="#1565c0"
        ).pack(pady=(10, 5))

        info = tk.Frame(scroll_frame)
        info.pack(fill="x", padx=10, pady=5)
        for label, val in [
            ("Всього позицій:", stats["total_products"]),
            ("Загальна сума:", f"{stats['total_sum']:,.2f} грн"),
            ("Середня ціна за шт:", f"{stats['avg_price']:,.2f} грн"),
            ("Загальна кількість одиниць:", f"{stats['total_quantity']:g}"),
        ]:
            row = tk.Frame(info)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=label, font=("Segoe UI", 10), anchor="w").pack(side="left")
            tk.Label(row, text=str(val), font=("Segoe UI", 10, "bold"), anchor="e").pack(
                side="right"
            )

        add_section("💰 Сума за матеріалами", stats.get("by_material", {}))
        add_section("📁 Сума за категоріями", stats.get("by_category", {}))

        tk.Button(scroll_frame, text="Закрити", command=self.destroy, width=15).pack(pady=15)
        self.wait_window(self)


class HistoryDialog(tk.Toplevel):
    def __init__(self, parent, product: Product):
        super().__init__(parent)
        self.title(f"📜 Історія цін — {product.name}")
        self.geometry("400x300")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        tk.Label(
            self,
            text=f"Історія зміни ціни: {product.name}",
            font=("Segoe UI", 11, "bold"),
            fg="#1565c0",
        ).pack(pady=10)

        cols = ("date", "old", "new", "diff")
        tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        tree.heading("date", text="Дата")
        tree.heading("old", text="Стара ціна")
        tree.heading("new", text="Нова ціна")
        tree.heading("diff", text="Різниця")

        tree.column("date", width=120, anchor="center")
        tree.column("old", width=80, anchor="e")
        tree.column("new", width=80, anchor="e")
        tree.column("diff", width=80, anchor="e")

        for h in reversed(product.price_history):
            diff = h.new_price - h.old_price
            tree.insert(
                "",
                "end",
                values=(h.date, f"{h.old_price:,.2f}", f"{h.new_price:,.2f}", f"{diff:+.2f}"),
            )

        tree.pack(fill="both", expand=True, padx=10, pady=5)
        tk.Button(self, text="Закрити", command=self.destroy, width=15).pack(pady=10)
        self.wait_window(self)


class ImportDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("📥 Імпорт")
        self.geometry("400x250")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.result = None

        tk.Label(
            self, text="Імпорт з CSV або Excel", font=("Segoe UI", 12, "bold"), fg="#1565c0"
        ).pack(pady=15)

        tk.Label(
            self,
            text="Формат файлів:\n"
            "CSV: Дата;Назва;Категорія;Довжина;Ширина;Висота;Діаметр;Матеріал;Товщина;К-ть;Ціна;Примітки\n"
            "Excel: аналогічно, з заголовком у першому рядку",
            font=("Segoe UI", 9),
            fg="#555",
            justify="center",
        ).pack(pady=5)

        self.skip_duplicates = tk.BooleanVar(value=True)
        tk.Checkbutton(
            self,
            text="Пропускати дублікати за назвою",
            variable=self.skip_duplicates,
            font=("Segoe UI", 9),
        ).pack(pady=5)

        frm = tk.Frame(self)
        frm.pack(pady=20)
        tk.Button(
            frm,
            text="📄 Імпортувати CSV",
            command=self._import_csv,
            bg="#1565c0",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            padx=15,
        ).pack(side="left", padx=5)
        if HAS_OPENPYXL:
            tk.Button(
                frm,
                text="📊 Імпортувати Excel",
                command=self._import_excel,
                bg="#2e7d32",
                fg="white",
                font=("Segoe UI", 10, "bold"),
                padx=15,
            ).pack(side="left", padx=5)
        tk.Button(frm, text="Скасувати", command=self.destroy, font=("Segoe UI", 10)).pack(
            side="left", padx=5
        )

        self.wait_window(self)

    def _parse_row(self, row: list) -> Product | None:
        try:
            if len(row) < 2:
                return None

            def get_val(idx, default=""):
                return str(row[idx]).strip() if len(row) > idx and row[idx] is not None else default

            def get_float(idx, default=0.0):
                val = row[idx] if len(row) > idx and row[idx] is not None else None
                if val == "" or val is None:
                    return default
                return float(val)

            date_added = get_val(0, datetime.now().strftime("%d.%m.%Y %H:%M"))
            name = get_val(1)
            if not name:
                return None
            category = get_val(2, "Інше")
            length = get_float(3)
            width = get_float(4)
            height = get_float(5)
            diameter = get_float(6)
            material = get_val(7)
            thickness = get_float(8)
            quantity = get_float(9, 1.0)
            price = get_float(10)
            notes = get_val(11)

            return Product(
                product_id=generate_id(),
                date_added=date_added,
                name=name,
                price_per_unit=price,
                quantity=quantity,
                length=length,
                width=width,
                height=height,
                diameter=diameter,
                material=material,
                thickness=thickness,
                category=category,
                notes=notes,
            )
        except Exception:
            return None

    def _import_csv(self):
        path = filedialog.askopenfilename(
            parent=self,
            title="Виберіть CSV-файл",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not path:
            return
        products = []
        existing_names = set()
        try:
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=";")
                next(reader, None)
                for row in reader:
                    p = self._parse_row(row)
                    if p:
                        if self.skip_duplicates.get() and p.name in existing_names:
                            continue
                        existing_names.add(p.name)
                        products.append(p)
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося імпортувати CSV:\n{e}")
            return
        self.result = products
        self.destroy()

    def _import_excel(self):
        if not HAS_OPENPYXL:
            return
        path = filedialog.askopenfilename(
            parent=self,
            title="Виберіть Excel-файл",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        products = []
        existing_names = set()
        try:
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                p = self._parse_row(list(row))
                if p:
                    if self.skip_duplicates.get() and p.name in existing_names:
                        continue
                    existing_names.add(p.name)
                    products.append(p)
        except Exception as e:
            messagebox.showerror("Помилка", f"Не вдалося імпортувати Excel:\n{e}")
            return
        self.result = products
        self.destroy()


class PriceListTab:
    def __init__(self, parent_frame: tk.Frame, root: tk.Tk, colors: dict | None = None):
        self.frame = parent_frame
        self.root = root
        self.colors = colors

        self.pl = PriceList()
        self._init_demo_data()

        self._sort_column = None
        self._sort_reverse = False

        self._headings = {
            "date_added": "Дата додавання",
            "name": "Назва виробу",
            "category": "Категорія",
            "length": "Довжина, мм",
            "width": "Ширина, мм",
            "height": "Висота, мм",
            "diameter": "Діаметр, мм",
            "material": "Матеріал",
            "thickness": "Товщина, мм",
            "quantity": "К-ть",
            "price_per_unit": "Ціна за шт, грн",
            "total_price": "Загальна, грн",
        }

        self.col_map = [
            ("date_added", "str", False, False, 110),
            ("name", "str", True, False, 180),
            ("category", "str", True, True, 120),
            ("length", "float", True, False, 75),
            ("width", "float", True, False, 75),
            ("height", "float", True, False, 75),
            ("diameter", "float", True, False, 80),
            ("material", "str", True, True, 100),
            ("thickness", "float", True, False, 80),
            ("quantity", "float", True, False, 55),
            ("price_per_unit", "float", True, False, 100),
            ("total_price", "float", False, False, 100),
        ]

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=28)
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#e3f2fd")

        toolbar = tk.Frame(self.frame, bg="#1565c0", height=50)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        btn_cfg = {
            "font": ("Segoe UI", 10, "bold"),
            "bg": "#1565c0",
            "fg": "white",
            "activebackground": "#0d47a1",
            "activeforeground": "white",
            "bd": 0,
            "padx": 12,
            "pady": 6,
            "cursor": "hand2",
        }

        tk.Button(toolbar, text="➕ Додати", command=self._add, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="🗑️ Видалити", command=self._delete, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="📋 Дублювати", command=self._duplicate, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="💰 Масова ціна", command=self._bulk_price, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="🖨️ Друк", command=self._print, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="📥 Імпорт", command=self._import, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="📊 Excel", command=self._export_excel, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="📜 Історія", command=self._show_history, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="📊 Статистика", command=self._show_statistics, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="↩️ Undo", command=self._undo, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="↪️ Redo", command=self._redo, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )
        tk.Button(toolbar, text="💾 Зберегти", command=self._save, **btn_cfg).pack(
            side="left", padx=5, pady=8
        )

        search_frame = tk.Frame(self.frame, bg="#f5f5f5", height=40)
        search_frame.pack(fill="x", padx=10, pady=(10, 0))
        search_frame.pack_propagate(False)
        tk.Label(search_frame, text="🔍 Пошук:", bg="#f5f5f5", font=("Segoe UI", 10)).pack(
            side="left", padx=5
        )
        self.search_var = tk.StringVar()
        self.search_var.trace("w", lambda *args: self._refresh_table())
        tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 10), width=40).pack(
            side="left", padx=5
        )
        tk.Label(
            search_frame,
            text="(назва, матеріал, категорія, примітки, дата)",
            bg="#f5f5f5",
            fg="#888",
            font=("Segoe UI", 9),
        ).pack(side="left", padx=5)

        table_frame = tk.Frame(self.frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        cols = [c[0] for c in self.col_map]
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")

        for col, (_attr, _typ, _editable, _is_combo, width) in enumerate(self.col_map):
            col_name = cols[col]
            self.tree.heading(
                col_name,
                text=self._headings.get(col_name, col_name),
                command=lambda c=col: self._sort_by(c),
            )
            self.tree.column(col_name, width=width, anchor="center" if col_name != "name" else "w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Button-1>", self._on_click_outside)
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Button-2>", self._show_context_menu)

        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="✏️ Редагувати", command=self._edit_selected)
        self.context_menu.add_command(label="📋 Дублювати", command=self._duplicate)
        self.context_menu.add_command(label="📜 Історія цін", command=self._show_history)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🗑️ Видалити", command=self._delete)

        self.sum_frame = tk.Frame(self.frame, bg="#e3f2fd", height=40)
        self.sum_frame.pack(fill="x", side="bottom")
        self.sum_frame.pack_propagate(False)
        self.sum_label = tk.Label(
            self.sum_frame,
            text="",
            font=("Segoe UI", 12, "bold"),
            bg="#e3f2fd",
            fg="#0d47a1",
            anchor="e",
        )
        self.sum_label.pack(fill="x", padx=15, pady=6)

        self.status = tk.Label(
            self.frame,
            text="Готово",
            bd=1,
            relief="sunken",
            anchor="w",
            font=("Segoe UI", 9),
            bg="#e3f2fd",
        )
        self.status.pack(fill="x", side="bottom")

        self._edit_widget = None
        self._edit_item = None
        self._edit_col_idx = None
        self._edit_attr = None
        self._edit_attr_type = None
        self._edit_is_combo = None

        self._autosave_enabled = False
        self._autosave_job = None
        self._dirty = False

        self._refresh_table()

    def _init_demo_data(self):
        if not self.pl.products:
            now = datetime.now().strftime("%d.%m.%Y %H:%M")
            self.pl.add(
                Product(
                    product_id=generate_id(),
                    date_added=now,
                    name="Вентилятор осьовий ВО-300",
                    price_per_unit=1250,
                    quantity=2,
                    length=0,
                    width=0,
                    height=0,
                    diameter=0,
                    material="цинк",
                    thickness=0,
                    category="Вентилятор",
                )
            )
            self.pl.add(
                Product(
                    product_id=generate_id(),
                    date_added=now,
                    name="Труба прямокутна ТП-100",
                    price_per_unit=450,
                    quantity=5,
                    length=1000,
                    width=100,
                    height=50,
                    diameter=0,
                    material="цинк",
                    thickness=0.5,
                    category="Труба прямокутна",
                )
            )
            self.pl.add(
                Product(
                    product_id=generate_id(),
                    date_added=now,
                    name="Труба кругла ТК-125",
                    price_per_unit=320,
                    quantity=10,
                    length=500,
                    width=125,
                    height=125,
                    diameter=125,
                    material="нержавійка",
                    thickness=0.4,
                    category="Труба кругла",
                )
            )
            self.pl.save()

    def _filtered_products(self) -> list[Product]:
        q = self.search_var.get().lower()
        if not q:
            return self.pl.products
        return [
            p
            for p in self.pl.products
            if q in p.name.lower()
            or q in p.material.lower()
            or q in p.category.lower()
            or q in p.notes.lower()
            or q in p.date_added.lower()
        ]

    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for p in self._filtered_products():
            self.tree.insert(
                "",
                "end",
                iid=p.id,
                values=(
                    p.date_added,
                    p.name,
                    p.category,
                    f"{p.length:g}" if p.length else "",
                    f"{p.width:g}" if p.width else "",
                    f"{p.height:g}" if p.height else "",
                    f"{p.diameter:g}" if p.diameter else "",
                    p.material,
                    f"{p.thickness:g}" if p.thickness else "",
                    f"{p.quantity:g}",
                    f"{p.price_per_unit:,.2f}",
                    f"{p.total_price:,.2f}",
                ),
            )

        total = len(self.pl.products)
        visible = len(self._filtered_products())
        grand = self.pl.total_sum()
        self.sum_label.config(text=f"💰 ЗАГАЛЬНА СУМА ПО ВСІХ ПОЗИЦІЯХ: {grand:,.2f} грн")
        self.status.config(
            text=f" Всього: {total} | Відображено: {visible} | "
            f"Подвійний клік — редагувати | Ctrl+N — новий | Ctrl+Z — скасувати"
        )

    def _selected_ids(self) -> list[str]:
        return list(self.tree.selection())

    def _sort_by(self, col_idx: int):
        attr_name = self.col_map[col_idx][0]
        if self._sort_column == attr_name:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_column = attr_name
            self._sort_reverse = False

        def sort_key(p: Product):
            val = getattr(p, attr_name, "")
            if isinstance(val, int | float):
                return (0, val)
            return (1, str(val).lower())

        self.pl.products.sort(key=sort_key, reverse=self._sort_reverse)
        self._refresh_table()

        arrow = " ▼" if self._sort_reverse else " ▲"
        for _col, (attr, _, _, _, _) in enumerate(self.col_map):
            text = self._headings.get(attr, attr)
            if attr == self._sort_column:
                text += arrow
            self.tree.heading(attr, text=text)

    def _on_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        if not item or not column:
            return

        col_idx = int(column[1:]) - 1
        attr_name, attr_type, editable, is_combobox, _ = self.col_map[col_idx]
        if not editable:
            return

        if self._edit_widget:
            self._save_inline()

        bbox = self.tree.bbox(item, column)
        if not bbox:
            return

        x, y, w, h = bbox
        current = self.tree.item(item, "values")[col_idx]

        if is_combobox:
            values = MATERIALS if attr_name == "material" else CATEGORIES
            self._edit_widget = ttk.Combobox(
                self.tree, values=values, state="readonly", font=("Segoe UI", 10)
            )
            self._edit_widget.set(current if current in values else values[0])
            self._edit_widget.place(x=x, y=y, width=w, height=h)
            self._edit_widget.focus()
            self._edit_widget.bind("<<ComboboxSelected>>", self._save_inline)
            self._edit_widget.bind("<Return>", self._save_inline)
            self._edit_widget.bind("<Escape>", self._cancel_inline)
        else:
            self._edit_widget = tk.Entry(
                self.tree, font=("Segoe UI", 10), justify="center" if attr_type != "str" else "left"
            )
            self._edit_widget.place(x=x, y=y, width=w, height=h)
            self._edit_widget.insert(0, current)
            self._edit_widget.select_range(0, tk.END)
            self._edit_widget.focus()
            self._edit_widget.bind("<Return>", self._save_inline)
            self._edit_widget.bind("<KP_Enter>", self._save_inline)
            self._edit_widget.bind("<Escape>", self._cancel_inline)

        self._edit_item = item
        self._edit_col_idx = col_idx
        self._edit_attr = attr_name
        self._edit_attr_type = attr_type
        self._edit_is_combo = is_combobox

    def _on_click_outside(self, event):
        if self._edit_widget:
            region = self.tree.identify("region", event.x, event.y)
            if region != "cell":
                self._save_inline()
            else:
                clicked_item = self.tree.identify_row(event.y)
                if clicked_item != self._edit_item:
                    self._save_inline()

    def _save_inline(self, event=None):
        if not self._edit_widget:
            return

        raw = self._edit_widget.get() if self._edit_is_combo else self._edit_widget.get().strip()

        self._edit_widget.destroy()
        self._edit_widget = None

        product = self.pl.get_by_id(self._edit_item)
        if not product:
            return

        if self._edit_attr_type == "float":
            try:
                new_val = float(raw) if raw else 0.0
            except ValueError:
                messagebox.showwarning("Увага", "Це поле має бути числом!")
                self._refresh_table()
                return
        else:
            new_val = raw

        if self._edit_attr == "price_per_unit" and new_val != product.price_per_unit:
            old_price = product.price_per_unit
            setattr(product, self._edit_attr, new_val)
            product.record_price_change(old_price)
        else:
            setattr(product, self._edit_attr, new_val)

        self._dirty = True
        self._refresh_table()
        self.pl.save()
        self.status.config(text=f" ✅ Збережено: {self._edit_attr} = {new_val}")

    def _cancel_inline(self, event=None):
        if self._edit_widget:
            self._edit_widget.destroy()
            self._edit_widget = None

    def _show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            if item not in self.tree.selection():
                self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def _edit_selected(self):
        self.status.config(text=" Використовуйте подвійний клік для редагування")

    def _add(self):
        dlg = AddProductDialog(self.root)
        if dlg.result:
            self.pl.add(dlg.result)
            self._dirty = True
            self._refresh_table()
            self.pl.save()
            self.status.config(text=f" ✅ Додано: {dlg.result.name}")

    def _duplicate(self):
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Увага", "Виберіть виріб для дублювання!")
            return
        if len(ids) > 1:
            messagebox.showwarning("Увага", "Виберіть лише один виріб для дублювання!")
            return
        original = self.pl.get_by_id(ids[0])
        if not original:
            return
        dlg = AddProductDialog(self.root, product_to_clone=original)
        if dlg.result:
            self.pl.add(dlg.result)
            self._dirty = True
            self._refresh_table()
            self.pl.save()
            self.status.config(text=f" ✅ Дубльовано: {dlg.result.name}")

    def _delete(self):
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Увага", "Виберіть виріб(и) для видалення!")
            return
        if messagebox.askyesno("Підтвердження", f"Видалити {len(ids)} виріб(ів)?"):
            for pid in ids:
                self.pl.delete(pid)
            self._dirty = True
            self._refresh_table()
            self.pl.save()

    def _bulk_price(self):
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Увага", "Виберіть виріб(и) для масового редагування!")
            return
        dlg = BulkPriceDialog(self.root, len(ids))
        if dlg.result:
            mode, val = dlg.result
            if mode == "abs":
                self.pl.bulk_edit_price(ids, val)
            else:
                self.pl.bulk_edit_price_percent(ids, val)
            self._dirty = True
            self._refresh_table()
            self.pl.save()

    def _show_history(self):
        ids = self._selected_ids()
        if not ids:
            messagebox.showwarning("Увага", "Виберіть виріб для перегляду історії!")
            return
        if len(ids) > 1:
            messagebox.showwarning("Увага", "Виберіть лише один виріб!")
            return
        product = self.pl.get_by_id(ids[0])
        if product:
            HistoryDialog(self.root, product)

    def _show_statistics(self):
        stats = self.pl.get_statistics()
        StatisticsDialog(self.root, stats)

    def _print(self):
        products = self._filtered_products()
        if not products:
            messagebox.showwarning("Увага", "Немає даних для друку!")
            return
        grand = self.pl.total_sum()
        html = generate_print_html(products, grand)
        fd, path = tempfile.mkstemp(suffix=".html", prefix="price_list_")
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file:///{path.replace(chr(92), '/')}")
        self.status.config(text=" Відкрито сторінку друку")

    def _export_excel(self):
        products = self._filtered_products()
        if not products:
            messagebox.showwarning("Увага", "Немає даних для експорту!")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror(
                "Бібліотека не встановлена",
                "Для експорту в Excel потрібна бібліотека openpyxl.\n\n"
                "Встановіть її командою:\n  pip install openpyxl",
            )
            return
        if not os.path.exists(EXPORT_DIR):
            os.makedirs(EXPORT_DIR)
        timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M")
        filename = f"price_list_{timestamp}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)
        grand = self.pl.total_sum()
        if export_to_excel(products, grand, filepath):
            self.status.config(text=f" 📊 Експортовано: {filepath}")
            if messagebox.askyesno("Готово", f"Файл збережено:\n{filepath}\n\nВідкрити папку?"):
                open_folder_crossplatform(EXPORT_DIR)
        else:
            messagebox.showerror("Помилка", "Не вдалося зберегти файл Excel.")

    def _import(self):
        dlg = ImportDialog(self.root)
        if dlg.result:
            self.pl.bulk_add(dlg.result)
            self._dirty = True
            self._refresh_table()
            self.pl.save()
            self.status.config(text=f" ✅ Імпортовано {len(dlg.result)} позицій")

    def _save(self):
        self.pl.save()
        self._dirty = False
        self.status.config(text=" 💾 Дані збережено у " + self.pl.filepath)

    def _undo(self):
        if self.pl.undo():
            self._dirty = True
            self._refresh_table()
            self.status.config(text=" ↩️ Скасовано останню дію")
        else:
            self.status.config(text=" Немає дій для скасування")

    def _redo(self):
        if self.pl.redo():
            self._dirty = True
            self._refresh_table()
            self.status.config(text=" ↪️ Повторено дію")
        else:
            self.status.config(text=" Немає дій для повторення")

    def _toggle_autosave(self):
        self._autosave_enabled = not self._autosave_enabled
        if self._autosave_enabled:
            self.status.config(text=" 🔄 Автозбереження увімкнено")
            self._schedule_autosave()
        else:
            self.status.config(text=" 🔄 Автозбереження вимкнено")
            if self._autosave_job:
                self.root.after_cancel(self._autosave_job)
                self._autosave_job = None

    def _schedule_autosave(self):
        if not self._autosave_enabled:
            return
        if self._dirty:
            self._save()
        self._autosave_job = self.root.after(300000, self._schedule_autosave)

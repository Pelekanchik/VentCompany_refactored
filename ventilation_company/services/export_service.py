"""
Сервіс експорту розрахунків (Excel, CSV, HTML, друк)
"""
import csv
import os
import tempfile
from typing import List, Dict, Any
from datetime import datetime


class ExportService:
    """Сервіс для експорту даних у різні формати."""

    @staticmethod
    def to_csv(data: List[Dict[str, Any]], filepath: str = None,
               headers: List[str] = None) -> str:
        """Експортує дані у CSV."""
        if filepath is None:
            filepath = os.path.join(tempfile.gettempdir(),
                                  f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

        if not data:
            return filepath

        if headers is None:
            headers = list(data[0].keys())

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)

        return filepath

    @staticmethod
    def to_excel(data: List[Dict[str, Any]], filepath: str = None,
                 sheet_name: str = "Розрахунок") -> str:
        """Експортує дані у Excel (якщо openpyxl встановлено)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            # Fallback на CSV
            return ExportService.to_csv(data, filepath.replace(".xlsx", ".csv") if filepath else None)

        if filepath is None:
            filepath = os.path.join(tempfile.gettempdir(),
                                  f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(filepath)
            return filepath

        headers = list(data[0].keys())

        # Заголовки
        header_fill = PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Дані
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        wb.save(filepath)
        return filepath

    @staticmethod
    def build_html_report(title: str, data: List[Dict[str, Any]],
                          summary: Dict[str, Any] = None) -> str:
        """Будує HTML-звіт для друку."""
        html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
body {{ font-family: "Segoe UI", Arial, sans-serif; margin: 40px; }}
h1 {{ color: #1565c0; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
th {{ background: #1565c0; color: white; }}
tr:nth-child(even) {{ background: #f5f5f5; }}
.summary {{ margin-top: 20px; padding: 15px; background: #e3f2fd; border-radius: 5px; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p>Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}</p>
"""

        if data:
            headers = list(data[0].keys())
            html += "<table><tr>"
            for h in headers:
                html += f"<th>{h}</th>"
            html += "</tr>"

            for row in data:
                html += "<tr>"
                for h in headers:
                    val = row.get(h, "")
                    if isinstance(val, float):
                        val = f"{val:,.2f}"
                    html += f"<td>{val}</td>"
                html += "</tr>"
            html += "</table>"

        if summary:
            html += '<div class="summary">'
            for key, val in summary.items():
                html += f"<p><b>{key}:</b> {val}</p>"
            html += "</div>"

        html += "</body></html>"
        return html

    @staticmethod
    def save_html_to_file(html_content: str, filepath: str = None) -> str:
        """Зберігає HTML у файл."""
        if filepath is None:
            filepath = os.path.join(tempfile.gettempdir(),
                                  f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)
        return filepath
